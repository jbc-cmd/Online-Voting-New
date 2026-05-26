"""
Exports Views — CSV, Excel, PDF export for election data.
"""
import csv
from datetime import datetime
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from elections.models import Election, Position
from students.models import Student
from voting.models import VoterParticipation, Vote
from audit.models import AuditLog, VerificationLog


# ─── CSV Exports ─────────────────────────────────────────────────────────────

@login_required
def export_results_csv(request, pk):
    election = get_object_or_404(Election, pk=pk)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="results_{election.slug}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Position', 'Candidate', 'Party List', 'Votes', 'Rank'])

    for position in election.positions.prefetch_related('candidates').all():
        candidates_data = []
        for candidate in position.candidates.all():
            vc = Vote.objects.filter(candidate=candidate, election=election, is_abstain=False).count()
            candidates_data.append((candidate, vc))
        candidates_data.sort(key=lambda x: x[1], reverse=True)
        for rank, (candidate, vc) in enumerate(candidates_data, 1):
            writer.writerow([position.title, candidate.name, candidate.partylist, vc, rank])
        abstain = Vote.objects.filter(position=position, election=election, is_abstain=True).count()
        if abstain > 0:
            writer.writerow([position.title, 'ABSTAIN', '', abstain, '-'])

    return response


@login_required
def export_turnout_csv(request, pk):
    election = get_object_or_404(Election, pk=pk)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="turnout_{election.slug}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Student ID', 'Full Name', 'Department', 'Program', 'Year Level', 'Section', 'Voted', 'Voted At', 'Ballot Number'])

    participations = VoterParticipation.objects.filter(
        election=election
    ).select_related('student', 'student__department', 'student__program')

    for p in participations:
        s = p.student
        writer.writerow([
            s.student_id_number,
            s.full_name,
            s.department.name if s.department else '',
            s.program.name if s.program else '',
            s.year_level,
            s.section,
            'Yes' if p.has_voted else 'No',
            p.voted_at.strftime('%Y-%m-%d %H:%M') if p.voted_at else '',
            p.ballot_number or '',
        ])

    return response


@login_required
def export_verification_logs_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="verification_logs.csv"'
    writer = csv.writer(response)
    writer.writerow(['DateTime', 'Election', 'Student ID Attempted', 'Email Attempted', 'Status', 'Reason', 'IP'])

    for log in VerificationLog.objects.select_related('election').order_by('-created_at')[:5000]:
        writer.writerow([
            log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            log.election.title if log.election else '',
            log.attempted_student_id_number,
            log.attempted_email,
            log.get_status_display(),
            log.reason,
            log.ip_address or '',
        ])
    return response


@login_required
def export_audit_logs_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="audit_logs.csv"'
    writer = csv.writer(response)
    writer.writerow(['DateTime', 'Actor', 'Action', 'Entity Type', 'Entity ID', 'IP'])

    for log in AuditLog.objects.select_related('actor').order_by('-created_at')[:5000]:
        writer.writerow([
            log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            log.actor.username if log.actor else 'System',
            log.action,
            log.entity_type,
            log.entity_id,
            log.ip_address or '',
        ])
    return response


# ─── Excel Exports ───────────────────────────────────────────────────────────

@login_required
def export_results_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    election = get_object_or_404(Election, pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Election Results'

    # Header
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1e3a5f')
    headers = ['Position', 'Candidate', 'Party List', 'Votes', 'Rank']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for position in election.positions.prefetch_related('candidates').all():
        candidates_data = []
        for candidate in position.candidates.all():
            vc = Vote.objects.filter(candidate=candidate, election=election, is_abstain=False).count()
            candidates_data.append((candidate, vc))
        candidates_data.sort(key=lambda x: x[1], reverse=True)
        for rank, (candidate, vc) in enumerate(candidates_data, 1):
            ws.cell(row=row_num, column=1, value=position.title)
            ws.cell(row=row_num, column=2, value=candidate.name)
            ws.cell(row=row_num, column=3, value=candidate.partylist)
            ws.cell(row=row_num, column=4, value=vc)
            ws.cell(row=row_num, column=5, value=rank)
            row_num += 1

    # Auto-width columns
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2.append(['Election', election.title])
    ws2.append(['Type', election.get_election_type_display()])
    ws2.append(['Status', election.get_status_display()])
    ws2.append(['Total Eligible Voters', election.total_eligible_voters])
    ws2.append(['Total Votes Cast', election.total_votes_cast])
    ws2.append(['Voter Turnout (%)', float(election.voter_turnout_percentage)])
    ws2.append(['Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="results_{election.slug}.xlsx"'
    wb.save(response)
    return response


# ─── PDF Exports ─────────────────────────────────────────────────────────────

@login_required
def export_results_pdf(request, pk):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    import io

    election = get_object_or_404(Election, pk=pk)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=12)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=13, spaceAfter=6)

    story = []
    story.append(Paragraph('UNIVOTE', title_style))
    story.append(Paragraph(f'Official Election Results Report', styles['Heading2']))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f'<b>Election:</b> {election.title}', styles['Normal']))
    story.append(Paragraph(f'<b>Type:</b> {election.get_election_type_display()}', styles['Normal']))
    story.append(Paragraph(f'<b>Status:</b> {election.get_status_display()}', styles['Normal']))
    story.append(Paragraph(f'<b>Total Eligible:</b> {election.total_eligible_voters}', styles['Normal']))
    story.append(Paragraph(f'<b>Total Voted:</b> {election.total_votes_cast}', styles['Normal']))
    story.append(Paragraph(f'<b>Turnout:</b> {election.voter_turnout_percentage}%', styles['Normal']))
    story.append(Paragraph(f'<b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
    story.append(Spacer(1, 1*cm))

    for position in election.positions.prefetch_related('candidates').all():
        story.append(Paragraph(position.title, heading_style))
        candidates_data = []
        for candidate in position.candidates.all():
            vc = Vote.objects.filter(candidate=candidate, election=election, is_abstain=False).count()
            candidates_data.append([candidate.name, candidate.partylist or '—', vc])
        candidates_data.sort(key=lambda x: x[2], reverse=True)

        table_data = [['Rank', 'Candidate', 'Party List', 'Votes']]
        for rank, row in enumerate(candidates_data, 1):
            table_data.append([rank] + row)

        abstain = Vote.objects.filter(position=position, election=election, is_abstain=True).count()
        if abstain > 0:
            table_data.append(['—', 'ABSTAIN', '', abstain])

        t = Table(table_data, colWidths=[2*cm, 7*cm, 5*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5*cm))

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="results_{election.slug}.pdf"'
    return response
