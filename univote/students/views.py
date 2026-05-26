"""
Students Views — Manage students, departments, programs, classes, organizations.
"""
import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages

from students.models import Student, Department, Program, ClassGroup, Organization
from audit.services import log_action_from_request


# ─── Students ────────────────────────────────────────────────────────────────

@login_required
def student_list(request):
    students = Student.objects.select_related(
        'department', 'program', 'class_group'
    ).all()

    # Filters
    dept_filter = request.GET.get('department')
    program_filter = request.GET.get('program')
    status_filter = request.GET.get('status')
    search = request.GET.get('search', '').strip()

    if dept_filter:
        students = students.filter(department_id=dept_filter)
    if program_filter:
        students = students.filter(program_id=program_filter)
    if status_filter:
        students = students.filter(status=status_filter)
    if search:
        students = students.filter(
            student_id_number__icontains=search
        ) | students.filter(
            first_name__icontains=search
        ) | students.filter(
            last_name__icontains=search
        ) | students.filter(
            official_school_email__icontains=search
        )

    departments = Department.objects.all()
    programs = Program.objects.all()

    return render(request, 'admin_panel/students/list.html', {
        'students': students,
        'departments': departments,
        'programs': programs,
        'dept_filter': dept_filter,
        'program_filter': program_filter,
        'status_filter': status_filter,
        'search': search,
    })


@login_required
def student_create(request):
    if request.method == 'POST':
        try:
            student = Student.objects.create(
                student_id_number=request.POST['student_id_number'].strip(),
                official_school_email=request.POST['official_school_email'].strip().lower(),
                first_name=request.POST['first_name'].strip(),
                last_name=request.POST['last_name'].strip(),
                department_id=request.POST.get('department') or None,
                program_id=request.POST.get('program') or None,
                year_level=request.POST.get('year_level', ''),
                section=request.POST.get('section', ''),
                class_group_id=request.POST.get('class_group') or None,
                status=request.POST.get('status', 'active'),
            )
            log_action_from_request(request, f'Created student: {student}', 'Student', student.pk)
            messages.success(request, 'Student added successfully.')
            return redirect('students:student_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin_panel/students/create.html', {
        'departments': Department.objects.all(),
        'programs': Program.objects.all(),
        'classes': ClassGroup.objects.all(),
    })


@login_required
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        try:
            student.student_id_number = request.POST['student_id_number'].strip()
            student.official_school_email = request.POST['official_school_email'].strip().lower()
            student.first_name = request.POST['first_name'].strip()
            student.last_name = request.POST['last_name'].strip()
            student.department_id = request.POST.get('department') or None
            student.program_id = request.POST.get('program') or None
            student.year_level = request.POST.get('year_level', '')
            student.section = request.POST.get('section', '')
            student.class_group_id = request.POST.get('class_group') or None
            student.status = request.POST.get('status', 'active')
            student.save()
            log_action_from_request(request, f'Updated student: {student}', 'Student', pk)
            messages.success(request, 'Student updated.')
            return redirect('students:student_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'admin_panel/students/edit.html', {
        'student': student,
        'departments': Department.objects.all(),
        'programs': Program.objects.all(),
        'classes': ClassGroup.objects.all(),
    })


@login_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        name = str(student)
        student.delete()
        log_action_from_request(request, f'Deleted student: {name}', 'Student', pk)
        messages.success(request, 'Student deleted.')
        return redirect('students:student_list')
    return render(request, 'admin_panel/students/delete_confirm.html', {'student': student})


@login_required
def student_import(request):
    """Import students from CSV or Excel."""
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, 'Please select a file.')
            return redirect('students:student_import')

        ext = file.name.split('.')[-1].lower()
        created = 0
        errors = []

        if ext == 'csv':
            try:
                decoded = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                for row in reader:
                    result = _import_student_row(row)
                    if result['success']:
                        created += 1
                    else:
                        errors.append(result['error'])
            except Exception as e:
                messages.error(request, f'CSV import error: {e}')
                return redirect('students:student_import')

        elif ext in ('xlsx', 'xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    result = _import_student_row(row_dict)
                    if result['success']:
                        created += 1
                    else:
                        errors.append(result['error'])
            except Exception as e:
                messages.error(request, f'Excel import error: {e}')
                return redirect('students:student_import')
        else:
            messages.error(request, 'Unsupported file type. Please use CSV or XLSX.')
            return redirect('students:student_import')

        log_action_from_request(request, f'Imported {created} students', 'Student')
        messages.success(request, f'Successfully imported {created} student(s). Errors: {len(errors)}.')
        if errors:
            for err in errors[:5]:
                messages.warning(request, err)
        return redirect('students:student_list')

    return render(request, 'admin_panel/students/import.html')


def _import_student_row(row: dict) -> dict:
    """Import a single student row. Returns {'success': bool, 'error': str}"""
    try:
        student_id = str(row.get('student_id_number', '')).strip()
        email = str(row.get('official_school_email', '')).strip().lower()
        first_name = str(row.get('first_name', '')).strip()
        last_name = str(row.get('last_name', '')).strip()

        if not all([student_id, email, first_name, last_name]):
            return {'success': False, 'error': f'Missing fields for ID: {student_id}'}

        dept_code = str(row.get('department_code', '')).strip()
        prog_code = str(row.get('program_code', '')).strip()

        dept = Department.objects.filter(code__iexact=dept_code).first()
        prog = Program.objects.filter(code__iexact=prog_code).first()
        year_level = str(row.get('year_level', '')).strip()
        section = str(row.get('section', '')).strip()

        class_group = None
        if prog and year_level and section:
            class_group = ClassGroup.objects.filter(
                program=prog, year_level=year_level, section__iexact=section
            ).first()

        Student.objects.update_or_create(
            student_id_number=student_id,
            defaults={
                'official_school_email': email,
                'first_name': first_name,
                'last_name': last_name,
                'department': dept,
                'program': prog,
                'year_level': year_level,
                'section': section,
                'class_group': class_group,
                'status': 'active',
            }
        )
        return {'success': True, 'error': None}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ─── Departments ─────────────────────────────────────────────────────────────

@login_required
def department_list(request):
    departments = Department.objects.prefetch_related('programs').all()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name', '').strip()
            code = request.POST.get('code', '').strip()
            if name and code:
                Department.objects.get_or_create(code=code, defaults={'name': name})
                messages.success(request, f'Department "{name}" added.')
        elif action == 'delete':
            Department.objects.filter(pk=request.POST.get('pk')).delete()
            messages.success(request, 'Department deleted.')
        return redirect('students:department_list')
    return render(request, 'admin_panel/departments/list.html', {'departments': departments})


# ─── Programs ────────────────────────────────────────────────────────────────

@login_required
def program_list(request):
    programs = Program.objects.select_related('department').all()
    departments = Department.objects.all()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name', '').strip()
            code = request.POST.get('code', '').strip()
            dept_pk = request.POST.get('department')
            if name and code and dept_pk:
                Program.objects.get_or_create(code=code, defaults={'name': name, 'department_id': dept_pk})
                messages.success(request, f'Program "{name}" added.')
        elif action == 'delete':
            Program.objects.filter(pk=request.POST.get('pk')).delete()
            messages.success(request, 'Program deleted.')
        return redirect('students:program_list')
    return render(request, 'admin_panel/programs/list.html', {
        'programs': programs, 'departments': departments
    })


# ─── Classes ─────────────────────────────────────────────────────────────────

@login_required
def class_list(request):
    classes = ClassGroup.objects.select_related('program').all()
    programs = Program.objects.all()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            prog_pk = request.POST.get('program')
            year_level = request.POST.get('year_level', '').strip()
            section = request.POST.get('section', '').strip()
            name = request.POST.get('name', '').strip()
            if prog_pk and year_level and section and name:
                ClassGroup.objects.get_or_create(
                    program_id=prog_pk, year_level=year_level, section=section,
                    defaults={'name': name}
                )
                messages.success(request, f'Class "{name}" added.')
        elif action == 'delete':
            ClassGroup.objects.filter(pk=request.POST.get('pk')).delete()
            messages.success(request, 'Class deleted.')
        return redirect('students:class_list')
    return render(request, 'admin_panel/classes/list.html', {
        'classes': classes, 'programs': programs,
        'year_levels': ClassGroup.YEAR_LEVEL_CHOICES,
    })


# ─── Organizations ───────────────────────────────────────────────────────────

@login_required
def organization_list(request):
    organizations = Organization.objects.all()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            if name:
                Organization.objects.get_or_create(name=name, defaults={'description': description})
                messages.success(request, f'Organization "{name}" added.')
        elif action == 'delete':
            Organization.objects.filter(pk=request.POST.get('pk')).delete()
            messages.success(request, 'Organization deleted.')
        return redirect('students:organization_list')
    return render(request, 'admin_panel/organizations/list.html', {'organizations': organizations})
